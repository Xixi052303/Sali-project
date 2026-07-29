#!/usr/bin/env python3
"""Apply preconditioned cell and table updates to an .xlsx file atomically."""

from __future__ import annotations

import argparse
import copy
import json
import os
import posixpath
import sys
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.worksheet.table import TableColumn


RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("payload", type=Path, help="UTF-8 JSON update payload")
    parser.add_argument("--dry-run", action="store_true", help="validate without saving")
    return parser.parse_args()


def _load_payload(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as stream:
        payload = json.load(stream)
    if not isinstance(payload, dict):
        raise ValueError("payload root must be an object")
    return payload


def _resolve_workbook(raw_path: Any) -> Path:
    if not isinstance(raw_path, str) or not raw_path.strip():
        raise ValueError("workbook must be a non-empty path string")
    workbook = Path(raw_path)
    if not workbook.is_absolute():
        workbook = Path.cwd() / workbook
    workbook = workbook.resolve()
    if workbook.suffix.lower() != ".xlsx":
        raise ValueError("only .xlsx workbooks are supported")
    if not workbook.is_file():
        raise FileNotFoundError(workbook)
    return workbook


def _short_repr(value: Any, limit: int = 160) -> str:
    rendered = ascii(value)
    if len(rendered) <= limit:
        return rendered
    return rendered[: limit - 3] + "..."


def _validate_expected_sheets(payload: dict[str, Any], sheet_names: list[str]) -> None:
    expected_sheets = payload.get("expected_sheets", [])
    if not isinstance(expected_sheets, list) or not all(isinstance(item, str) for item in expected_sheets):
        raise ValueError("expected_sheets must be a list of strings")
    missing = [name for name in expected_sheets if name not in sheet_names]
    if missing:
        raise ValueError(f"missing expected sheets: {missing}")


def _validate_updates(payload: dict[str, Any], sheet_names: list[str]) -> list[dict[str, Any]]:
    updates = payload.get("updates", [])
    if not isinstance(updates, list):
        raise ValueError("updates must be a list")

    seen: set[tuple[str, str]] = set()
    for index, update in enumerate(updates, start=1):
        if not isinstance(update, dict):
            raise ValueError(f"update {index} must be an object")
        for required in ("sheet", "cell", "expect", "value"):
            if required not in update:
                raise ValueError(f"update {index} missing {required!r}")
        sheet = update["sheet"]
        cell = update["cell"]
        if not isinstance(sheet, str) or sheet not in sheet_names:
            raise ValueError(f"update {index} has unknown sheet {sheet!r}")
        if not isinstance(cell, str):
            raise ValueError(f"update {index} cell must be a string")
        coordinate_to_tuple(cell)
        normalized_cell = cell.upper()
        target = (sheet, normalized_cell)
        if target in seen:
            raise ValueError(f"duplicate target {sheet}!{normalized_cell}")
        seen.add(target)
        kind = update.get("kind", "text")
        if kind not in {"text", "scalar", "formula"}:
            raise ValueError(f"update {index} has invalid kind {kind!r}")
        value = update["value"]
        if kind == "text" and not isinstance(value, str):
            raise ValueError(f"update {index} text value must be a string")
        if kind == "scalar" and value is not None and not isinstance(value, (bool, int, float)):
            raise ValueError(f"update {index} scalar value must be a number, boolean, or null")
        if kind == "formula" and (not isinstance(value, str) or not value.startswith("=")):
            raise ValueError(f"update {index} formula value must start with '='")
    return updates


def _find_table(workbook: Any, sheet_name: str, table_name: str) -> Any:
    table = workbook[sheet_name].tables.get(table_name)
    if table is None:
        raise ValueError(f"unknown table {sheet_name}!{table_name}")
    return table


def _validate_column_names(names: Any, label: str) -> list[str]:
    if not isinstance(names, list) or not names or not all(isinstance(item, str) for item in names):
        raise ValueError(f"{label} must be a non-empty list of strings")
    if any(not item.strip() for item in names):
        raise ValueError(f"{label} must not contain blank names")
    normalized = [item.casefold() for item in names]
    if len(normalized) != len(set(normalized)):
        raise ValueError(f"{label} must not contain duplicate names")
    return names


def _validate_table_updates(payload: dict[str, Any], workbook: Any) -> list[dict[str, Any]]:
    table_updates = payload.get("table_updates", [])
    if not isinstance(table_updates, list):
        raise ValueError("table_updates must be a list")

    seen: set[tuple[str, str]] = set()
    for index, update in enumerate(table_updates, start=1):
        if not isinstance(update, dict):
            raise ValueError(f"table update {index} must be an object")
        for required in (
            "sheet",
            "table",
            "expect_ref",
            "ref",
            "expect_columns",
            "expect_headers",
            "columns",
        ):
            if required not in update:
                raise ValueError(f"table update {index} missing {required!r}")

        sheet_name = update["sheet"]
        table_name = update["table"]
        if not isinstance(sheet_name, str) or sheet_name not in workbook.sheetnames:
            raise ValueError(f"table update {index} has unknown sheet {sheet_name!r}")
        if not isinstance(table_name, str) or not table_name:
            raise ValueError(f"table update {index} table must be a non-empty string")
        target = (sheet_name, table_name.casefold())
        if target in seen:
            raise ValueError(f"duplicate table update {sheet_name}!{table_name}")
        seen.add(target)

        table = _find_table(workbook, sheet_name, table_name)
        expect_ref = update["expect_ref"]
        new_ref = update["ref"]
        if not isinstance(expect_ref, str) or not isinstance(new_ref, str):
            raise ValueError(f"table update {index} refs must be strings")
        old_bounds = range_boundaries(expect_ref)
        new_bounds = range_boundaries(new_ref)
        if old_bounds[:2] != new_bounds[:2]:
            raise ValueError(f"table update {index} may not move the table start cell")
        if table.ref != expect_ref:
            raise ValueError(
                f"table precondition mismatch {sheet_name}!{table_name}: "
                f"expected ref {expect_ref!r}, found {table.ref!r}"
            )

        expect_columns = _validate_column_names(
            update["expect_columns"], f"table update {index} expect_columns"
        )
        columns = _validate_column_names(update["columns"], f"table update {index} columns")
        actual_columns = [column.name for column in table.tableColumns]
        if actual_columns != expect_columns:
            raise ValueError(
                f"table precondition mismatch {sheet_name}!{table_name}: "
                f"expected columns {_short_repr(expect_columns)}, found {_short_repr(actual_columns)}"
            )
        old_width = old_bounds[2] - old_bounds[0] + 1
        new_width = new_bounds[2] - new_bounds[0] + 1
        if len(expect_columns) != old_width:
            raise ValueError(f"table update {index} expect_columns width does not match expect_ref")
        if len(columns) != new_width:
            raise ValueError(f"table update {index} columns width does not match ref")

        expect_headers = update["expect_headers"]
        if not isinstance(expect_headers, list) or len(expect_headers) != new_width:
            raise ValueError(f"table update {index} expect_headers width does not match ref")
        worksheet = workbook[sheet_name]
        current_headers = [
            worksheet.cell(new_bounds[1], column_index).value
            for column_index in range(new_bounds[0], new_bounds[2] + 1)
        ]
        if current_headers != expect_headers:
            raise ValueError(
                f"table header precondition mismatch {sheet_name}!{table_name}: "
                f"expected {_short_repr(expect_headers)}, found {_short_repr(current_headers)}"
            )
    return table_updates


def _validate_table_package(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        members = set(archive.namelist())
        table_parts = {
            name for name in members if name.startswith("xl/tables/") and name.endswith(".xml")
        }
        referenced_parts: set[str] = set()
        for rel_path in members:
            if not rel_path.startswith("xl/worksheets/_rels/") or not rel_path.endswith(".rels"):
                continue
            source_name = rel_path.replace("/_rels/", "/")[: -len(".rels")]
            source_dir = posixpath.dirname(source_name)
            root = ET.fromstring(archive.read(rel_path))
            for relationship in root.findall(f"{{{RELATIONSHIP_NAMESPACE}}}Relationship"):
                if not relationship.attrib.get("Type", "").endswith("/table"):
                    continue
                if relationship.attrib.get("TargetMode") == "External":
                    raise ValueError(f"external table relationship is not supported: {rel_path}")
                target = relationship.attrib.get("Target", "")
                resolved = (
                    target.lstrip("/")
                    if target.startswith("/")
                    else posixpath.normpath(posixpath.join(source_dir, target))
                )
                if resolved not in members:
                    raise ValueError(f"missing table relationship target: {resolved}")
                referenced_parts.add(resolved)
        orphan_parts = sorted(table_parts - referenced_parts)
        if orphan_parts:
            raise ValueError(f"unreferenced table parts: {orphan_parts}")


def _ranges_overlap(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    return not (
        left[2] < right[0]
        or right[2] < left[0]
        or left[3] < right[1]
        or right[3] < left[1]
    )


# openpyxl 能重开文件并不代表 Excel 会接受表对象，因此显式核对表的跨部件约束。
def _validate_table_integrity(
    workbook: Any, ignored_tables: set[tuple[str, str]] | None = None
) -> None:
    ignored = ignored_tables or set()
    seen_ids: set[int] = set()
    seen_names: set[str] = set()
    for worksheet in workbook.worksheets:
        sheet_ranges: list[tuple[str, tuple[int, int, int, int]]] = []
        for table in worksheet.tables.values():
            table_key = (worksheet.title, table.displayName.casefold())
            if table_key in ignored:
                continue
            try:
                bounds = range_boundaries(table.ref)
            except ValueError as error:
                raise ValueError(
                    f"invalid table ref {worksheet.title}!{table.displayName}: {table.ref!r}"
                ) from error
            width = bounds[2] - bounds[0] + 1
            columns = list(table.tableColumns)
            if len(columns) != width:
                raise ValueError(
                    f"table column count mismatch {worksheet.title}!{table.displayName}: "
                    f"ref width {width}, columns {len(columns)}"
                )
            header_values = [
                worksheet.cell(bounds[1], column_index).value
                for column_index in range(bounds[0], bounds[2] + 1)
            ]
            column_names = [column.name for column in columns]
            normalized_headers = [None if value is None else str(value) for value in header_values]
            if normalized_headers != column_names:
                raise ValueError(
                    f"table header mismatch {worksheet.title}!{table.displayName}: "
                    f"cells {_short_repr(normalized_headers)}, columns {_short_repr(column_names)}"
                )
            _validate_column_names(column_names, f"table {worksheet.title}!{table.displayName} columns")
            if table.id in seen_ids:
                raise ValueError(f"duplicate table id: {table.id}")
            seen_ids.add(table.id)
            normalized_name = table.displayName.casefold()
            if normalized_name in seen_names:
                raise ValueError(f"duplicate table name: {table.displayName}")
            seen_names.add(normalized_name)
            if table.autoFilter is not None and table.autoFilter.ref not in (None, table.ref):
                raise ValueError(
                    f"table autoFilter mismatch {worksheet.title}!{table.displayName}: "
                    f"{table.autoFilter.ref!r} != {table.ref!r}"
                )
            for other_name, other_bounds in sheet_ranges:
                if _ranges_overlap(bounds, other_bounds):
                    raise ValueError(
                        f"overlapping tables on {worksheet.title}: {other_name} and {table.displayName}"
                    )
            sheet_ranges.append((table.displayName, bounds))


def _validate_structural_update_targets(workbook: Any, updates: list[dict[str, Any]]) -> None:
    violations: list[str] = []
    for update in updates:
        worksheet = workbook[update["sheet"]]
        row_index, column_index = coordinate_to_tuple(update["cell"])
        for table in worksheet.tables.values():
            min_column, min_row, max_column, _ = range_boundaries(table.ref)
            if row_index != min_row:
                continue
            if min_column <= column_index <= max_column:
                violations.append(
                    f"{worksheet.title}!{update['cell']} is a header of table {table.displayName}"
                )
            elif column_index in (min_column - 1, max_column + 1) and update["value"] not in (None, ""):
                violations.append(
                    f"{worksheet.title}!{update['cell']} is adjacent to table {table.displayName} "
                    "and may be an undeclared table expansion"
                )
    if violations:
        raise ValueError(
            "cell updates may not change table structure; use table_updates:\n" + "\n".join(violations)
        )


def _apply_updates(workbook: Any, updates: list[dict[str, Any]]) -> None:
    mismatches: list[str] = []
    for update in updates:
        cell = workbook[update["sheet"]][update["cell"]]
        if cell.value != update["expect"]:
            mismatches.append(
                f"{ascii(update['sheet'])}!{update['cell']}: "
                f"expected {_short_repr(update['expect'])}, found {_short_repr(cell.value)}"
            )
    if mismatches:
        raise ValueError("precondition mismatch:\n" + "\n".join(mismatches))

    for update in updates:
        cell = workbook[update["sheet"]][update["cell"]]
        cell.value = update["value"]
        if update.get("kind", "text") == "text" and update["value"].startswith("="):
            cell.data_type = "s"


# 表结构更新必须同步工作表表头和 tableColumns，避免 Excel 自动修复 table*.xml。
def _apply_table_updates(workbook: Any, table_updates: list[dict[str, Any]]) -> None:
    for update in table_updates:
        worksheet = workbook[update["sheet"]]
        table = _find_table(workbook, update["sheet"], update["table"])
        old_columns = list(table.tableColumns)
        new_columns: list[TableColumn] = []
        for index, name in enumerate(update["columns"], start=1):
            if index <= len(old_columns):
                column = copy.copy(old_columns[index - 1])
                column.id = index
                column.name = name
            else:
                column = TableColumn(id=index, name=name)
            new_columns.append(column)
        table.tableColumns = new_columns
        table.ref = update["ref"]
        if table.autoFilter is not None:
            table.autoFilter.ref = update["ref"]
        min_column, min_row, max_column, _ = range_boundaries(update["ref"])
        for column_index, name in zip(range(min_column, max_column + 1), update["columns"]):
            worksheet.cell(min_row, column_index).value = name


def _verify_table_updates(workbook: Any, table_updates: list[dict[str, Any]]) -> None:
    mismatches: list[str] = []
    for update in table_updates:
        table = _find_table(workbook, update["sheet"], update["table"])
        actual_columns = [column.name for column in table.tableColumns]
        if table.ref != update["ref"] or actual_columns != update["columns"]:
            mismatches.append(
                f"{update['sheet']}!{update['table']}: ref {table.ref!r}, "
                f"columns {_short_repr(actual_columns)}"
            )
    if mismatches:
        raise ValueError("saved table mismatch:\n" + "\n".join(mismatches))


def _verify_saved(
    path: Path,
    expected_sheets: list[str],
    updates: list[dict[str, Any]],
    table_updates: list[dict[str, Any]],
) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        broken_member = archive.testzip()
        if broken_member is not None:
            raise ValueError(f"invalid workbook archive member: {broken_member}")
    _validate_table_package(path)
    workbook = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    try:
        if workbook.sheetnames != expected_sheets:
            raise ValueError("sheet order changed during save")
        _validate_table_integrity(workbook)
        mismatches = []
        for update in updates:
            actual = workbook[update["sheet"]][update["cell"]].value
            if actual != update["value"]:
                mismatches.append(
                    f"{ascii(update['sheet'])}!{update['cell']}: saved {_short_repr(actual)}"
                )
        if mismatches:
            raise ValueError("saved value mismatch:\n" + "\n".join(mismatches))
        _verify_table_updates(workbook, table_updates)
    finally:
        workbook.close()


def main() -> int:
    args = _parse_args()
    payload = _load_payload(args.payload.resolve())
    workbook_path = _resolve_workbook(payload.get("workbook"))
    original_stat = workbook_path.stat()
    workbook = load_workbook(workbook_path, data_only=False, read_only=False, keep_links=True)
    temp_path: Path | None = None
    try:
        original_sheets = list(workbook.sheetnames)
        _validate_expected_sheets(payload, original_sheets)
        updates = _validate_updates(payload, original_sheets)
        table_updates = _validate_table_updates(payload, workbook)
        if not updates and not table_updates:
            raise ValueError("payload must contain at least one cell or table update")

        ignored_tables = {
            (update["sheet"], update["table"].casefold()) for update in table_updates
        }
        _validate_table_integrity(workbook, ignored_tables)
        _validate_structural_update_targets(workbook, updates)
        _apply_updates(workbook, updates)
        _apply_table_updates(workbook, table_updates)
        _validate_table_integrity(workbook)
        if args.dry_run:
            print(
                json.dumps(
                    {
                        "status": "dry-run-ok",
                        "workbook": str(workbook_path),
                        "updates": len(updates),
                        "table_updates": len(table_updates),
                    }
                )
            )
            return 0

        handle, raw_temp_path = tempfile.mkstemp(
            prefix=f".{workbook_path.stem}.", suffix=".xlsx", dir=workbook_path.parent
        )
        os.close(handle)
        temp_path = Path(raw_temp_path)
        workbook.save(temp_path)
        _verify_saved(temp_path, original_sheets, updates, table_updates)

        current_stat = workbook_path.stat()
        if (current_stat.st_mtime_ns, current_stat.st_size) != (original_stat.st_mtime_ns, original_stat.st_size):
            raise RuntimeError("source workbook changed during update; refusing to replace it")
        os.replace(temp_path, workbook_path)
        temp_path = None
        _verify_saved(workbook_path, original_sheets, updates, table_updates)
        print(
            json.dumps(
                {
                    "status": "saved",
                    "workbook": str(workbook_path),
                    "updates": len(updates),
                    "table_updates": len(table_updates),
                }
            )
        )
        return 0
    finally:
        workbook.close()
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(json.dumps({"status": "error", "error": str(error)}), file=sys.stderr)
        raise SystemExit(1)
