from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_DIR = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPT_DIR))

import patch_workbook  # noqa: E402


class PatchWorkbookTableTests(unittest.TestCase):
    def _create_workbook(self, directory: str) -> Path:
        path = Path(directory) / "table.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "事件"
        worksheet.append(["请求时间(s)", "事件ID"])
        worksheet.append([1, "start"])
        worksheet.append([2, "boss"])
        table = Table(displayName="TimelineEventsTable", ref="A1:B3")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        workbook.save(path)
        workbook.close()
        return path

    def test_valid_table_and_package_pass(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = self._create_workbook(directory)
            workbook = load_workbook(path, data_only=False, read_only=False)
            try:
                patch_workbook._validate_table_integrity(workbook)
                patch_workbook._validate_table_package(path)
            finally:
                workbook.close()

    def test_header_drift_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = self._create_workbook(directory)
            workbook = load_workbook(path, data_only=False, read_only=False)
            try:
                workbook["事件"]["A1"] = "路程进度"
                with self.assertRaisesRegex(ValueError, "table header mismatch"):
                    patch_workbook._validate_table_integrity(workbook)
            finally:
                workbook.close()

    def test_cell_update_cannot_change_table_header(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = self._create_workbook(directory)
            workbook = load_workbook(path, data_only=False, read_only=False)
            try:
                updates = [
                    {
                        "sheet": "事件",
                        "cell": "A1",
                        "expect": "请求时间(s)",
                        "value": "路程进度",
                        "kind": "text",
                    }
                ]
                with self.assertRaisesRegex(ValueError, "use table_updates"):
                    patch_workbook._validate_structural_update_targets(workbook, updates)
            finally:
                workbook.close()

    def test_explicit_table_update_repairs_header_and_expands(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = self._create_workbook(directory)
            workbook = load_workbook(path, data_only=False, read_only=False)
            try:
                worksheet = workbook["事件"]
                worksheet["A1"] = "路程进度"
                worksheet["C1"] = "前进倍率"
                payload = {
                    "table_updates": [
                        {
                            "sheet": "事件",
                            "table": "TimelineEventsTable",
                            "expect_ref": "A1:B3",
                            "ref": "A1:C3",
                            "expect_columns": ["请求时间(s)", "事件ID"],
                            "expect_headers": ["路程进度", "事件ID", "前进倍率"],
                            "columns": ["路程进度", "事件ID", "前进倍率"],
                        }
                    ]
                }
                table_updates = patch_workbook._validate_table_updates(payload, workbook)
                patch_workbook._apply_table_updates(workbook, table_updates)
                patch_workbook._validate_table_integrity(workbook)
                table = worksheet.tables["TimelineEventsTable"]
                self.assertEqual(table.ref, "A1:C3")
                self.assertEqual(
                    [column.name for column in table.tableColumns],
                    ["路程进度", "事件ID", "前进倍率"],
                )
            finally:
                workbook.close()

    def test_normal_cell_update_still_saves_and_verifies(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = self._create_workbook(directory)
            workbook = load_workbook(path, data_only=False, read_only=False)
            updates = [
                {
                    "sheet": "事件",
                    "cell": "B2",
                    "expect": "start",
                    "value": "intro",
                    "kind": "text",
                }
            ]
            try:
                patch_workbook._validate_structural_update_targets(workbook, updates)
                patch_workbook._apply_updates(workbook, updates)
                workbook.save(path)
            finally:
                workbook.close()
            patch_workbook._verify_saved(path, ["事件"], updates, [])


if __name__ == "__main__":
    unittest.main()
